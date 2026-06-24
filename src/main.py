import os
import sys
import toml
import yaml  # Для чтения индивидуальных config.yaml отчетов
import sqlite3
import psycopg2
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

now = datetime.now()
# === КОСТЫЛЬ СОВМЕСТИМОСТИ ДЛЯ PARAMIKO & SSHTUNNEL ===
import paramiko

if not hasattr(paramiko, 'DSSKey'):
    class DummyDSSKey: pass
    paramiko.DSSKey = DummyDSSKey

from sshtunnel import SSHTunnelForwarder

# 1. Получаем абсолютный путь к папке, где физически лежит этот main.py
_current_dir = os.path.dirname(os.path.abspath(__file__))

# 2. Если main.py лежит в src, корень — на уровень выше. Иначе — это и есть корень.
if os.path.basename(_current_dir) == "src":
    BASE_DIR = os.path.dirname(_current_dir)
else:
    BASE_DIR = _current_dir

# 3. Абсолютные пути к ключевым точкам системы
MAIN_CONFIG_PATH = os.path.join(BASE_DIR, "config", "main.toml")
JOBS_DIR = os.path.join(BASE_DIR, "jobs")

# print(f"[DEBUG PATHS] Корень проекта определен как: {BASE_DIR}")
# print(f"[DEBUG PATHS] Центральный конфиг ищется по: {MAIN_CONFIG_PATH}")
# print(f"[DEBUG PATHS] Папка с отчетами находится по: {JOBS_DIR}")


def get_connection_for_job(job_name):
    """
    Универсальная фабрика подключений.
    Динамически ищет конфигурацию БД в main.toml по ключу из секции [jobs].
    Никакого хардкода структуры секций или типов.
    """
    if not os.path.exists(MAIN_CONFIG_PATH):
        raise FileNotFoundError(f"Критическая ошибка: Главный конфиг не найден: {MAIN_CONFIG_PATH}")

    config = toml.load(MAIN_CONFIG_PATH)
    jobs_section = config.get("jobs", {})
    db_profile_key = jobs_section.get(job_name)

    if not db_profile_key:
        raise ValueError(f"Отчет '{job_name}' не зарегистрирован в секции [jobs] в main.toml")

    # --- ДИНАМИЧЕСКИЙ ПОИСК ПРОФИЛЯ ПО КЛЮЧУ ---
    # Поддерживает как "database.postgres_prod", так и вложенные структуры любой глубины
    db_config = config
    for part in db_profile_key.split('.'):
        if isinstance(db_config, dict):
            db_config = db_config.get(part)
        else:
            db_config = None
            break

    if not db_config or not isinstance(db_config, dict):
        raise ValueError(f"Конфигурация подключения [{db_profile_key}] не найдена или некорректна в main.toml")
    # --------------------------------------------

    # Определяем тип СУБД строго из параметров самого профиля в TOML
    db_type = db_config.get("type")

    # 1. Режим СУБД: SQLite
    if db_type == "sqlite":
        db_path = db_config.get("database")
        print(f"[ENGINE] Подключение к локальной СУБД SQLite: {db_path}")
        return sqlite3.connect(db_path), None

    # 2. Режим СУБД: PostgreSQL
    if db_type == "postgres":
        if db_config.get("use_ssh", False):
            print(f"[ENGINE] Инициализация SSH-туннеля для профиля [{db_profile_key}]...")
            try:
                tunnel_kwargs = {
                    "ssh_address_or_host": (db_config.get("ssh_host"), int(db_config.get("ssh_port", 22))),
                    "ssh_username": db_config.get("ssh_user"),
                    "remote_bind_address": (db_config.get("host"), int(db_config.get("port", 5432)))
                }

                # Загрузка учетных данных SSH из TOML
                if db_config.get("ssh_password"):
                    tunnel_kwargs["ssh_password"] = db_config.get("ssh_password")
                elif db_config.get("ssh_pkey"):
                    clean_key_path = os.path.normpath(db_config.get("ssh_pkey"))
                    if os.path.exists(clean_key_path):
                        try:
                            tunnel_kwargs["ssh_pkey"] = paramiko.RSAKey.from_private_key_file(clean_key_path)
                        except Exception:
                            tunnel_kwargs["ssh_pkey"] = paramiko.Ed25519Key.from_private_key_file(clean_key_path)
                    else:
                        raise FileNotFoundError(f"Файл SSH-ключа не найден: {clean_key_path}")

                tunnel = SSHTunnelForwarder(**tunnel_kwargs)
                tunnel.start()
                print(f"[ENGINE] SSH-туннель успешно поднят на локальном порту: {tunnel.local_bind_port}")

                # Подключение к Postgres через локальную точку туннеля
                conn = psycopg2.connect(
                    host='127.0.0.1',
                    port=tunnel.local_bind_port,
                    user=db_config.get("user"),
                    password=db_config.get("password"),
                    database=db_config.get("database"),
                    connect_timeout=10
                )
                return conn, tunnel
            except Exception as e:
                print(f"[💥 ENGINE ERROR] Сбой SSH-туннелирования: {e}")
                raise e
        else:
            # Прямое подключение к Postgres (без SSH) на основе параметров TOML
            target_host = db_config.get("host")
            target_port = int(db_config.get("port", 5432))
            print(f"[ENGINE] Прямое подключение к СУБД Postgres ({target_host}:{target_port})...")
            conn = psycopg2.connect(
                host=target_host,
                port=target_port,
                user=db_config.get("user"),
                password=db_config.get("password"),
                database=db_config.get("database"),
                connect_timeout=10
            )
            return conn, None

    raise ValueError(f"Неподдерживаемый тип СУБД '{db_type}' в профиле конфигурации [{db_profile_key}]")


def get_all_jobs():
    """ Динамический список задач из центрального TOML """
    try:
        config = toml.load(MAIN_CONFIG_PATH)
        return list(config.get("jobs", {}).keys())
    except Exception:
        return []


def handle_delivery(job_config, file_path):
    """
    Блок дистрибуции отчета на основе секции delivery в индивидуальном config.yaml
    """
    delivery = job_config.get("delivery", {})
    if not delivery:
        return

    # Локальное сохранение/архивирование
    local_cfg = delivery.get("local", {})
    if local_cfg.get("enabled"):
        target_dir = local_cfg.get("target_path")
        if target_dir:
            try:
                os.makedirs(target_dir, exist_ok=True)
                import shutil
                shutil.copy(file_path, os.path.join(target_dir, os.path.basename(file_path)))
                print(f"[DELIVERY] Файл успешно скопирован в локальный архив: {target_dir}")
            except Exception as e:
                print(f"[💥 DELIVERY ERROR] Локальное архивирование сорвалось: {e}")

    # Интеграция с внешними сервисами доставки при необходимости настраивается здесь
    if delivery.get("mail", {}).get("enabled"):
        print("[DELIVERY] Внешняя рассылка Mail активна (параметры из main.toml)...")

    # Внешняя выгрузка в Nextcloud (с поддержкой нескольких профилей)
    nc_config = delivery.get("nextcloud", {})
    if nc_config.get("enabled"):
        print("[DELIVERY] Запуск выгрузки отчета в Nextcloud...")

        # 1. Загружаем глобальные доступы из центрального конфига main.toml
        try:
            import toml
            global_config = toml.load(MAIN_CONFIG_PATH)
            # Извлекаем всю секцию nextcloud со всеми профилями
            nc_delivery_all = global_config.get("delivery", {}).get("nextcloud", {})
        except Exception as e:
            print(f"[💥 NEXTCLOUD ERROR] Не удалось прочитать main.toml: {e}")
            return

        # Определяем, какой профиль затребован отчетом (по умолчанию covid)
        profile_name = nc_config.get("profile", "covid")
        # Берем настройки конкретного профиля (например, 'covid') из прочитанного словаря
        nc_global = nc_delivery_all.get(profile_name, {})

        if not nc_global:
            print(f"[💥 NEXTCLOUD ERROR] Профиль '{profile_name}' не найден в конфигурации main.toml!")
            return

        server_url = nc_global.get("server_url", "").rstrip("/")
        username = nc_global.get("username")
        password = nc_global.get("password")
        path_user = nc_global.get("path_user", username)

        remote_path = nc_config.get("remote_path", "").strip("/")
        file_name = os.path.basename(file_path)

        if not server_url or not username or not password:
            print(f"[💥 NEXTCLOUD ERROR] В профиле '{profile_name}' отсутствуют настройки подключения")
            return

        # 2. Формируем WebDAV URL с безопасным кодированием кириллицы и пробелов
        import urllib.parse

        if remote_path:
            encoded_path = "/".join([urllib.parse.quote(part) for part in remote_path.split("/")])
            webdav_url = f"{server_url}/remote.php/dav/files/{path_user}/{encoded_path}/{urllib.parse.quote(file_name)}"
        else:
            webdav_url = f"{server_url}/remote.php/dav/files/{path_user}/{urllib.parse.quote(file_name)}"

        # 3. Отправляем файл методом PUT с игнорированием корпоративных SSL
        try:
            import requests
            import urllib3
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

            print(f"[DELIVERY] Отправка файла в профиль '{profile_name}'...")
            with open(file_path, "rb") as f:
                response = requests.put(
                    webdav_url,
                    auth=(username, password),
                    data=f,
                    verify=False,
                    timeout=60
                )

            if response.status_code in [201, 204]:
                path_log = f"{remote_path}/{file_name}" if remote_path else file_name
                print(f"[✅ NEXTCLOUD SUCCESS] [{profile_name}] Файл успешно загружен: {path_log}")
            else:
                print(
                    f"[💥 NEXTCLOUD ERROR] [{profile_name}] Ошибка загрузки (Код: {response.status_code}): {response.text}")

        except Exception as n_ex:
            print(f"[💥 NEXTCLOUD CRITICAL ERROR] [{profile_name}] Ошибка при передаче данных: {n_ex}")


def run_job(job_name, user_params=None):
    if user_params is None:
        user_params = {}

    print(f"\n[WORKER] Инициализация конвейера отчета: {job_name}")

    # Строим пути строго на основе глобальной JOBS_DIR
    job_dir = os.path.join(JOBS_DIR, job_name)
    yaml_path = os.path.join(job_dir, "config.yaml")

    if not os.path.exists(yaml_path):
        raise FileNotFoundError(f"Критическая ошибка: Конфигурационный файл {yaml_path} не найден!")

    # 2. Загрузка конфигурации отчета
    with open(yaml_path, "r", encoding="utf-8") as f:
        job_config = yaml.safe_load(f)

    if not job_config.get("enabled", True):
        print(f"[WORKER] Генерация отчета '{job_name}' отменена: статус 'enabled: false'")
        return None

    # --- УМНЫЙ И БЕЗОПАСНЫЙ СБОР ПАРАМЕТРОВ ---
    defined_params = job_config.get("parameters", {}) or {}
    final_params = {}

    # Заполняем дефолтными значениями из config.yaml
    for p_name, p_info in defined_params.items():
        if isinstance(p_info, dict):
            final_params[p_name] = p_info.get("default")

    # Накатываем то, что пришло от пользователя/бота поверх дефолтов
    if user_params and isinstance(user_params, dict):
        for k, v in user_params.items():
            final_params[k] = v

    # КОНВЕРТАЦИЯ СТРОКОВЫХ МАКРОСОВ В РЕАЛЬНЫЕ ДАТЫ ДЛЯ СУБД
    from datetime import timedelta
    now = datetime.now()

    for k, v in final_params.items():
        if isinstance(v, str):
            # 1. ГЛУБОКАЯ ОЧИСТКА ВВОДА ОТ МУСОРА ИЗ EXCEL
            # Удаляем переводы строк, возвраты каретки и табуляцию, заменяя их на пустоту
            v = v.replace("\n", "").replace("\r", "").replace("\t", "")

            # Заменяем скрытые неразрывные пробелы (\xa0) на обычные
            v = v.replace("\xa0", " ")

            # Удаляем любые кавычки (одинарные, двойные, обратные апострофы, французские ёлочки)
            for quote in ["`", "'", '"', "«", "»"]:
                v = v.replace(quote, "")

            # Срезаем лишние пробелы по краям, которые могли остаться
            v = v.strip()

            # Пересохраняем очищенное значение обратно в словарь
            final_params[k] = v

            # 2. ДАЛЬШЕ ИДЕТ ВАШ СТАНДАРТНЫЙ БЛОК ПАРСИНГА МАКРОСОВ
            val_clean = v.lower()

            if val_clean == "today":
                final_params[k] = now.date()
            val_clean = v.lower().strip()
            if val_clean == "today":
                final_params[k] = now.date()
            elif val_clean == "minus_7_days":
                final_params[k] = (now - timedelta(days=7)).date()
            elif val_clean == "minus_30_days":
                final_params[k] = (now - timedelta(days=30)).date()

    print(f"[ENGINE] Итоговый набор параметров для СУБД (после парсинга макросов): {final_params}")
    # --------------------------------------------------------------

    # === ИСПРАВЛЕНО: Старое ошибочное чтение несуществующего sql_file отсюда УБРАНО ===

    workbook_cfg = job_config.get("workbook", {})
    sheets_cfg = workbook_cfg.get("sheets", [])

    if not sheets_cfg:
        raise ValueError(f"Конфигурация '{job_name}' не содержит описания листов книги (sheets)")

    # Подключаемся к базе, используя полностью динамический парсер профилей
    conn, tunnel = get_connection_for_job(job_name)
    cursor = conn.cursor()

    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Удаляем дефолтный лист

        # Стилизация книги
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"),
                             top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))

        # Обработка каждой вкладки отчета последовательно
        for sheet_idx, sheet_info in enumerate(sheets_cfg):
            sheet_name = sheet_info.get("name", f"Лист {sheet_idx + 1}")
            sql_file_name = sheet_info.get("sql_file")
            columns_mapping = sheet_info.get("columns", {}) or {}

            # ЧИТАЕМ SQL-ФАЙЛ СТРОГО ЗДЕСЬ — ВНУТРИ ЦИКЛА ВКЛАДКИ
            sql_path = os.path.join(job_dir, sql_file_name)
            if not os.path.exists(sql_path):
                raise FileNotFoundError(f"Не найден файл запроса {sql_path} для листа '{sheet_name}'")

            with open(sql_path, "r", encoding="utf-8") as sf:
                sql_query = sf.read()

            print(f"[ENGINE] Сбор данных для листа '{sheet_name}' (SQL: {sql_file_name})...")

            # ИСПРАВЛЕНО: Передаем отвалидированный final_params вместо сырого user_params
            cursor.execute(sql_query, final_params)
            rows = cursor.fetchall()

            # ... Дальше идет ваш стандартный код заполнения Excel (ws = wb.create_sheet...)

            # Валидация на пустые выборки на основе правил из config.yaml
            validation = job_config.get("validation", {})
            if len(rows) == 0 and not validation.get("allow_empty", True):
                if validation.get("on_empty_action") == "alert":
                    raise ValueError(
                        f"Критический сбой валидации данных: Вкладка '{sheet_name}' пуста. Выполнение прервано.")

            # Создание целевой вкладки
            ws = wb.create_sheet(title=sheet_name)
            ws.views.sheetView[0].showGridLines = True

            # Читаем оригинальные системные имена столбцов, возвращенные СУБД
            db_columns = [desc[0] for desc in cursor.description]
            # Маппим оригинальные имена колонок в пользовательские русские имена из config.yaml
            headers = [columns_mapping.get(col, col) for col in db_columns]

            # Добавление и стилизация шапки таблицы
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            # Заполнение данными
            for row_data in rows:
                clean_row = []
                for item in row_data:
                    # 1. Если это дата и время — форматируем в красивую строку
                    if isinstance(item, datetime):
                        clean_row.append(item.strftime("%Y-%m-%d %H:%M:%S"))
                    # 2. Если поле бинарное (bytes или memoryview) — пишем заглушку
                    elif isinstance(item, (bytes, memoryview)):
                        clean_row.append("бинарный код")
                    # 3. Все остальные стандартные типы (числа, строки, None) оставляем как есть
                    else:
                        clean_row.append(item)

                ws.append(clean_row)
                current_row = ws.max_row
                for col_num in range(1, len(clean_row) + 1):
                    ws.cell(row=current_row, column=col_num).border = thin_border

            # Корректировка ширины столбцов под размер контента
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            ws.freeze_panes = "A2"

        # Сборка финального имени файла по маске filename_template
        template = workbook_cfg.get("filename_template", f"{job_name}_{{YYYY}}_{{MM}}_{{DD}}.xlsx")
        now = datetime.now()
        filename = template.format(
            YYYY=now.strftime("%Y"),
            MM=now.strftime("%m"),
            DD=now.strftime("%d")
        )

        output_file_path = os.path.join(os.getcwd(), "output", filename)
        os.makedirs(os.path.dirname(output_file_path), exist_ok=True)

        wb.save(output_file_path)
        wb.close()
        print(f"[EXCEL] Книга Excel успешно сформирована: {output_file_path}")

        # Локальная или облачная дистрибуция готового файла
        handle_delivery(job_config, output_file_path)

        return os.path.abspath(output_file_path)

    except Exception as err:
        print(f"[💥 WORKER ERROR] Ошибка генерации отчета '{job_name}': {err}")
        raise err
    finally:
        cursor.close()
        conn.close()
        if tunnel:
            tunnel.stop()
            print("[ENGINE] Сессия SSH-туннеля закрыта.")


if __name__ == "__main__":
    print("=== Запуск движка отчетов с динамической фабрикой СУБД ===")
    try:
        # Тест с передачей параметров (если они требуются вашим .sql скриптам)
        path = run_job("pharmacy", user_params={"status_id": 1})
        print(f"[УСПЕХ] Готовый файл находится здесь:\n{path}")
    except Exception as e:
        print(f"[ОШИБКА]: {e}")